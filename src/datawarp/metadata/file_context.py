"""
File context extraction for two-stage enrichment.

Stage 0: Extract raw text from metadata sheets (Notes, Contents, Definitions)
Stage 1: LLM extracts structured context (one call per file)
"""
import json
import os
from dataclasses import dataclass, field
from typing import Dict, List, Optional

# Sheets that contain metadata (not data)
METADATA_PATTERNS = ('contents', 'index', 'toc', 'notes', 'methodology', 'definitions', 'glossary',
                     'data source', 'cover', 'about', 'title', 'key facts', 'introduction')


@dataclass
class FileContext:
    """Structured context extracted from file metadata sheets."""
    sheets: Dict[str, str] = field(default_factory=dict)   # sheet_name → description
    kpis: Dict[str, str] = field(default_factory=dict)     # kpi_name → definition
    definitions: Dict[str, str] = field(default_factory=dict)  # measure → precise clinical definition
    methodology: str = ""
    data_sources: List[str] = field(default_factory=list)
    codes: Dict[str, str] = field(default_factory=dict)    # code → meaning (SNOMED, ICD, etc.)

    def to_dict(self) -> dict:
        return {'sheets': self.sheets, 'kpis': self.kpis, 'definitions': self.definitions,
                'methodology': self.methodology, 'data_sources': self.data_sources, 'codes': self.codes}

    @classmethod
    def from_dict(cls, data: dict) -> 'FileContext':
        return cls(sheets=data.get('sheets', {}), kpis=data.get('kpis', {}),
                   definitions=data.get('definitions', {}), methodology=data.get('methodology', ''),
                   data_sources=data.get('data_sources', []), codes=data.get('codes', {}))


def extract_metadata_text(file_path: str, max_rows: int = 50) -> str:
    """Stage 0: Extract raw text from all metadata sheets in file."""
    if not file_path.endswith(('.xlsx', '.xls')):
        return ""

    try:
        import openpyxl
        from datawarp.loader import get_sheet_names

        sheets = get_sheet_names(file_path)
        metadata_sheets = [s for s in sheets if any(p in s.lower() for p in METADATA_PATTERNS)]
        if not metadata_sheets:
            return ""

        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        sections = []

        for sheet_name in metadata_sheets:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            lines = []
            for i, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True)):
                if i >= max_rows:
                    break
                vals = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if vals:
                    lines.append(' | '.join(vals))
            if lines:
                sections.append(f"=== {sheet_name} ===\n" + '\n'.join(lines))

        wb.close()
        return '\n\n'.join(sections)
    except Exception:
        return ""


def extract_file_context(metadata_text: str, all_sheets: List[str] = None, pipeline_id: str = "", source_file: str = "") -> Optional[FileContext]:
    """Stage 1: Call LLM once to extract structured context from metadata text."""
    if not metadata_text or len(metadata_text) < 50:
        return None

    try:
        from litellm import completion, completion_cost
    except ImportError:
        return None

    import time
    from .enrich import _log_enrichment_call

    provider = os.getenv('LLM_PROVIDER', 'gemini')
    model = os.getenv('LLM_MODEL', 'gemini-2.0-flash')
    model_id = f"{provider}/{model}" if provider == 'gemini' else model

    sheets_hint = f"\nAll sheets: {all_sheets}" if all_sheets else ""
    prompt = f"""Extract structured metadata from NHS file documentation.
{sheets_hint}

{metadata_text[:6000]}

Return JSON with these fields:
- sheets: Map sheet names to their descriptions
- kpis: High-level KPI names and what they measure
- definitions: PRECISE clinical definitions (e.g., "smoker = current smoker +/-3 days of delivery", timing, thresholds, inclusion criteria)
- methodology: Data collection notes, rounding rules, suppression thresholds
- data_sources: Source datasets (e.g., "MSDS", "HES")
- codes: Any SNOMED, ICD, or classification codes mentioned with their meanings

IMPORTANT: For definitions, extract the EXACT clinical criteria - timing windows, thresholds, what counts as a "case". These are critical for data interpretation.

Return JSON only:
{{"sheets": {{}}, "kpis": {{}}, "definitions": {{}}, "methodology": "", "data_sources": [], "codes": {{}}}}"""

    start_time = time.time()
    log_data = {'pipeline_id': pipeline_id, 'source_file': source_file, 'sheet_name': '[FILE_CONTEXT]',
                'provider': provider, 'model': model, 'prompt_text': prompt}

    try:
        response = completion(model=model_id, messages=[{"role": "user", "content": prompt}],
                              max_tokens=1500, temperature=0.1, timeout=30)
        text = response.choices[0].message.content
        log_data['duration_ms'] = int((time.time() - start_time) * 1000)
        log_data['response_text'] = text

        # Extract usage
        usage = getattr(response, 'usage', None)
        if usage:
            log_data['input_tokens'] = getattr(usage, 'prompt_tokens', 0)
            log_data['output_tokens'] = getattr(usage, 'completion_tokens', 0)
            log_data['total_tokens'] = getattr(usage, 'total_tokens', 0)
        try:
            log_data['cost_usd'] = completion_cost(completion_response=response)
        except Exception:
            pass

        if "```" in text:
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]

        result = FileContext.from_dict(json.loads(text.strip()))
        log_data['success'] = True
        log_data['suggested_table_name'] = f"context:{len(result.sheets)}sheets,{len(result.kpis)}kpis"
        _log_enrichment_call(log_data)
        return result
    except Exception as e:
        log_data['duration_ms'] = int((time.time() - start_time) * 1000)
        log_data['success'] = False
        log_data['error_message'] = str(e)
        _log_enrichment_call(log_data)
        print(f"Warning: File context extraction failed: {e}")
        return None
