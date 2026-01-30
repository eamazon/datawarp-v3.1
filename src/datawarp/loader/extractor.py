"""
FileExtractor - Ported from DataWarp v3 with optimizations.

Key features:
- Multi-row header detection
- Hierarchical column names from merged cells
- Type inference using Excel cell metadata
- Footer/duplicate header detection

CRITICAL: This extracts structure. The loader uses DataFrame.columns as single source of truth.
"""

import openpyxl
from openpyxl.utils import get_column_letter
import re
import logging
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field
from enum import Enum, auto
from pathlib import Path

logger = logging.getLogger(__name__)

# Workbook cache: filepath → openpyxl.Workbook
_workbook_cache: Dict[str, Any] = {}


def clear_workbook_cache():
    """Clear the workbook cache. Call at end of batch processing."""
    for wb in _workbook_cache.values():
        try:
            wb.close()
        except Exception:
            pass
    _workbook_cache.clear()


def _get_cached_workbook(filepath: str):
    """Get workbook from cache or load and cache it."""
    if filepath not in _workbook_cache:
        logger.debug(f"Loading workbook: {Path(filepath).name}")
        _workbook_cache[filepath] = openpyxl.load_workbook(filepath, data_only=True)
    return _workbook_cache[filepath]


class SheetType(Enum):
    TABULAR = auto()
    METADATA = auto()
    EMPTY = auto()
    UNRECOGNISED = auto()


@dataclass
class ColumnInfo:
    """Column metadata from Excel extraction."""
    excel_col: str
    col_index: int
    pg_name: str
    original_headers: List[str]
    inferred_type: str = 'VARCHAR(255)'
    sample_values: List[Any] = field(default_factory=list)

    @property
    def full_header(self) -> str:
        return ' > '.join(h for h in self.original_headers if h)


@dataclass
class TableStructure:
    """Detected table structure from Excel sheet."""
    sheet_name: str
    sheet_type: SheetType
    header_rows: List[int]
    data_start_row: int
    data_end_row: int
    columns: Dict[int, ColumnInfo]
    error_message: Optional[str] = None

    @property
    def is_valid(self) -> bool:
        return self.sheet_type == SheetType.TABULAR and self.error_message is None

    @property
    def total_data_rows(self) -> int:
        return max(0, self.data_end_row - self.data_start_row + 1)

    def get_column_names(self) -> List[str]:
        return [self.columns[idx].pg_name for idx in sorted(self.columns.keys())]


class FileExtractor:
    """
    Optimized extractor for NHS England Excel publications.

    Row-major cell access pattern for performance.
    """

    PATTERNS = {
        'fiscal_year': re.compile(r'^\d{4}[-/]\d{2,4}$'),
        'calendar_year': re.compile(r'^(19|20)\d{2}$'),
        'month_year': re.compile(r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s*\d{4}$', re.I),
        'quarter': re.compile(r'^Q?[1-4]$', re.I),
        'org_code': re.compile(r'^[A-Z0-9]{2,5}$'),
        'fy_header': re.compile(r'^FY\s*\d{4}', re.I),
    }

    STOP_WORDS = ('note', 'source', 'copyright', '©', 'please', 'this worksheet', 'this table')
    SUPPRESSED_VALUES = {':', '..', '.', '-', '*', 'c', 'z', 'x', '[c]', '[z]', '[x]', 'n/a', 'na'}
    METADATA_INDICATORS = ('contents', 'title', 'notes', 'definition', 'about', 'introduction')

    def __init__(self, filepath: str, sheet_name: str):
        import warnings
        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

        self.filepath = Path(filepath)
        self.wb = _get_cached_workbook(filepath)

        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {self.wb.sheetnames}")

        self.sheet_name = sheet_name
        self.ws = self.wb[self.sheet_name]
        self._structure: Optional[TableStructure] = None
        self._merged_map: Dict[Tuple[int, int], Tuple[int, int, str]] = {}
        self._row_cache: Dict[int, List[Any]] = {}

        self._build_merged_map()

    def _build_merged_map(self):
        """Build map of merged cell ranges."""
        for mr in self.ws.merged_cells.ranges:
            val = self.ws.cell(row=mr.min_row, column=mr.min_col).value
            val_str = str(val).replace('\n', ' ').strip() if val else ""
            for row in range(mr.min_row, mr.max_row + 1):
                for col in range(mr.min_col, mr.max_col + 1):
                    self._merged_map[(row, col)] = (mr.min_row, mr.min_col, val_str)

    def _cache_rows(self, rows: List[int], max_col: int):
        """Pre-cache multiple rows in row-major order."""
        for row in rows:
            if row not in self._row_cache:
                self._row_cache[row] = [
                    self.ws.cell(row=row, column=col).value
                    for col in range(1, max_col + 1)
                ]

    def _get_cached_value(self, row: int, col: int) -> Any:
        """Get value from cache (col is 1-indexed)."""
        if row in self._row_cache and col <= len(self._row_cache[row]):
            return self._row_cache[row][col - 1]
        return self.ws.cell(row=row, column=col).value

    def _get_cached_value_str(self, row: int, col: int) -> str:
        """Get string value from cache, handling merged cells."""
        if (row, col) in self._merged_map:
            return self._merged_map[(row, col)][2]

        val = self._get_cached_value(row, col)
        return str(val).replace('\n', ' ').strip() if val else ""

    def _get_cell_value(self, row: int, col: int) -> str:
        """Get cell value, resolving merged cells."""
        if (row, col) in self._merged_map:
            return self._merged_map[(row, col)][2]
        val = self.ws.cell(row=row, column=col).value
        return str(val).replace('\n', ' ').strip() if val else ""

    def infer_structure(self) -> TableStructure:
        """Auto-detect complete table structure."""
        if self._structure:
            return self._structure

        sheet_type = self._classify_sheet()

        if sheet_type != SheetType.TABULAR:
            self._structure = TableStructure(
                sheet_name=self.sheet_name,
                sheet_type=sheet_type,
                header_rows=[],
                data_start_row=0,
                data_end_row=0,
                columns={},
            )
            return self._structure

        try:
            header_rows = self._detect_all_header_rows()

            if not header_rows:
                raise ValueError("Could not detect header rows")

            data_start_row = max(header_rows) + 1
            while data_start_row <= min(self.ws.max_row, max(header_rows) + 5):
                if self._is_data_row(data_start_row):
                    break
                data_start_row += 1

            columns = self._build_column_hierarchy(header_rows, data_start_row)

            if not columns:
                raise ValueError("No data columns detected")

            data_end_row = self._find_data_end(data_start_row)

            self._infer_column_types(columns, data_start_row, data_end_row)

            self._structure = TableStructure(
                sheet_name=self.sheet_name,
                sheet_type=SheetType.TABULAR,
                header_rows=header_rows,
                data_start_row=data_start_row,
                data_end_row=data_end_row,
                columns=columns,
            )

        except Exception as e:
            self._structure = TableStructure(
                sheet_name=self.sheet_name,
                sheet_type=SheetType.UNRECOGNISED,
                header_rows=[],
                data_start_row=0,
                data_end_row=0,
                columns={},
                error_message=str(e)
            )

        return self._structure

    def _classify_sheet(self) -> SheetType:
        """Classify sheet type with smart tabular detection.

        Uses data density and row structure analysis to distinguish real
        tabular data from documentation/metadata sheets (like NHS data definitions).
        """
        if self.ws.max_row < 2 or self.ws.max_column < 2:
            return SheetType.EMPTY

        # Analyze first 30 rows for density and structure
        empty, single, multi, total_cells = 0, 0, 0, 0
        for row in range(1, min(31, self.ws.max_row + 1)):
            cells = sum(1 for col in range(1, min(20, self.ws.max_column + 1))
                        if self.ws.cell(row=row, column=col).value is not None)
            total_cells += cells
            if cells == 0:
                empty += 1
            elif cells <= 2:
                single += 1
            else:
                multi += 1

        # Calculate metrics
        sample_rows = min(30, self.ws.max_row)
        density = total_cells / (sample_rows * 10) if sample_rows > 0 else 0
        single_ratio = single / max(1, single + multi)

        # Low density (<15%) OR high single-cell ratio (>50%) with few multi-cell rows = documentation
        if density < 0.15 or (single_ratio > 0.5 and multi < 5):
            return SheetType.METADATA

        # Check first cell for metadata indicators
        first_val = self.ws.cell(row=1, column=1).value
        if first_val:
            val_lower = str(first_val).lower()
            if any(ind in val_lower for ind in self.METADATA_INDICATORS):
                if multi < 3:
                    return SheetType.METADATA

        return SheetType.TABULAR if multi >= 3 else SheetType.METADATA

    def _detect_all_header_rows(self) -> List[int]:
        """Detect header rows - handles multi-row headers."""
        header_rows = []

        merge_rows = set()
        for mr in self.ws.merged_cells.ranges:
            if mr.max_col - mr.min_col >= 1:
                merge_rows.add(mr.min_row)

        first_header_row = None

        for row_num in range(1, 30):
            cells = self._count_cells(row_num)

            if cells < 2:
                continue

            val_a = self._get_cell_value(row_num, 1)
            val_b = self._get_cell_value(row_num, 2)

            if val_a.lower().startswith(('table', 'this', 'note')):
                continue

            val_b_stripped = val_b.strip()
            if val_b_stripped and val_b_stripped.endswith(':') and len(val_b_stripped) < 25:
                if not val_a.strip():
                    continue

            real_numeric_count = 0
            text_count = 0
            year_count = 0
            period_count = 0
            unit_count = 0

            for col in range(1, min(20, self.ws.max_column + 1)):
                val = self._get_cell_value(row_num, col)
                if not val:
                    continue

                val_clean = val.strip()

                if val_clean.lower() in self.SUPPRESSED_VALUES or val_clean == ':':
                    continue

                if self._is_unit_label(val_clean):
                    unit_count += 1
                    continue

                if self.PATTERNS['calendar_year'].match(val_clean):
                    year_count += 1
                    continue

                if self.PATTERNS['fiscal_year'].match(val_clean.rstrip('²³¹')):
                    year_count += 1
                    continue

                if re.match(r'^[QH][1-4]$', val_clean, re.I):
                    period_count += 1
                    continue

                if self._is_real_numeric_data(val_clean):
                    real_numeric_count += 1
                else:
                    text_count += 1

            is_data_row = real_numeric_count >= 2

            is_header_row = (
                row_num in merge_rows or
                (year_count >= 2 and real_numeric_count == 0) or
                period_count >= 2 or
                unit_count >= 2 or
                (real_numeric_count == 0 and text_count >= 2)
            )

            if first_header_row is None and is_header_row and (cells >= 2 or row_num in merge_rows):
                first_header_row = row_num

            if first_header_row is not None:
                if is_data_row and not is_header_row:
                    header_rows = list(range(first_header_row, row_num))
                    break

        if not header_rows and first_header_row:
            header_rows = [first_header_row]

        return header_rows

    def _find_max_column(self, header_rows: List[int], data_start_row: int) -> int:
        """Find max column in headers and data."""
        max_col = 1

        rows_to_check = list(header_rows) + list(range(data_start_row, min(data_start_row + 5, self.ws.max_row + 1)))

        for row in rows_to_check:
            for col in range(min(self.ws.max_column, 500), 0, -1):
                val = self.ws.cell(row=row, column=col).value
                if val is not None:
                    max_col = max(max_col, col)
                    break

        return max_col

    def _build_column_hierarchy(
        self,
        header_rows: List[int],
        data_start_row: int
    ) -> Dict[int, ColumnInfo]:
        """Build column metadata from multi-row headers."""
        columns = {}

        max_col = self._find_max_column(header_rows, data_start_row)

        rows_to_cache = list(header_rows) + list(range(
            data_start_row,
            min(data_start_row + 10, self.ws.max_row + 1)
        ))
        self._cache_rows(rows_to_cache, max_col)

        used_names = {}
        has_data_rows = list(range(data_start_row, min(data_start_row + 5, self.ws.max_row + 1)))

        for col in range(1, max_col + 1):
            header_values = [self._get_cached_value_str(row, col) for row in header_rows]

            has_data = any(
                self._get_cached_value(r, col) is not None
                for r in has_data_rows
            )

            all_headers_empty = all(not h for h in header_values)

            if not has_data and all_headers_empty:
                continue  # Skip spacer columns

            # Build unique headers
            unique_headers = []
            for h in header_values:
                if h and (not unique_headers or h != unique_headers[-1]):
                    unique_headers.append(h)

            if unique_headers:
                raw_name = '_'.join(unique_headers)
            else:
                raw_name = f"column_{get_column_letter(col)}"

            pg_name = self._to_db_identifier(raw_name)

            if pg_name in used_names:
                used_names[pg_name] += 1
                suffix = f"_{used_names[pg_name]}"
                if len(pg_name) + len(suffix) > 63:
                    pg_name = pg_name[:63 - len(suffix)]
                pg_name = f"{pg_name}{suffix}"
            else:
                used_names[pg_name] = 0

            sample_rows = list(range(data_start_row, min(data_start_row + 10, self.ws.max_row + 1)))
            samples = [self._get_cached_value(r, col) for r in sample_rows]

            col_info = ColumnInfo(
                excel_col=get_column_letter(col),
                col_index=col,
                pg_name=pg_name,
                original_headers=header_values,
                sample_values=samples
            )

            columns[col] = col_info

        return columns

    def _infer_column_types(
        self,
        columns: Dict[int, ColumnInfo],
        data_start_row: int,
        data_end_row: int
    ):
        """Infer column types using Excel cell metadata."""
        for col_idx, col_info in columns.items():
            cell_types_seen = set()
            sample_values = []
            has_decimal_values = False

            for r in range(data_start_row, data_end_row + 1):
                cell = self.ws.cell(row=r, column=col_idx)
                val = cell.value

                if val is not None:
                    if cell.data_type:
                        cell_types_seen.add(cell.data_type)

                        if cell.data_type == 'n' and isinstance(val, (int, float)):
                            if val % 1 != 0:
                                has_decimal_values = True
                    else:
                        if isinstance(val, str):
                            cell_types_seen.add('s')

                if r < data_start_row + 100:
                    sample_values.append(val)

            col_info.sample_values = sample_values

            has_numeric = 'n' in cell_types_seen or 'd' in cell_types_seen
            has_text = 's' in cell_types_seen

            if has_numeric and has_text:
                col_info.inferred_type = 'VARCHAR(255)'
            elif has_numeric and has_decimal_values:
                col_info.inferred_type = 'DOUBLE PRECISION'
            else:
                col_info.inferred_type = self._infer_type_from_values(
                    col_info.sample_values,
                    col_info.pg_name
                )

    def _infer_type_from_values(self, values: List[Any], col_name: str) -> str:
        """Infer type from sample values."""
        has_suppression = any(
            str(v).strip().lower() in self.SUPPRESSED_VALUES
            for v in values if v is not None
        )

        if has_suppression:
            return 'VARCHAR(255)'

        clean = [v for v in values if v is not None and str(v).strip().lower() not in self.SUPPRESSED_VALUES]

        if not clean:
            return 'VARCHAR(255)'

        name_lower = col_name.lower()

        if any(x in name_lower for x in ['date', 'month', 'year', 'quarter', 'period']):
            return 'VARCHAR(255)'
        if any(x in name_lower for x in ['description', 'definition', 'notes', 'comment', 'detail']):
            return 'TEXT'  # Descriptions can be very long
        if any(x in name_lower for x in ['name', 'category', 'group', 'trust', 'type']):
            return 'VARCHAR(255)'
        if any(x in name_lower for x in ['code', 'org', 'ics', 'nhse']):
            return 'VARCHAR(20)'

        int_count = 0
        float_count = 0
        text_count = 0

        for val in clean[:25]:
            s = str(val).strip()
            if ' - ' in s or ' to ' in s.lower():
                text_count += 1
                continue
            try:
                float(s.replace(',', '').replace('£', '').replace('$', '').replace('%', ''))
                if '.' in s:
                    float_count += 1
                else:
                    int_count += 1
            except ValueError:
                text_count += 1

        total = len(clean[:25])

        if total == 0:
            return 'VARCHAR(255)'

        if int_count / total > 0.7 and float_count == 0:
            return 'INTEGER'

        if (int_count + float_count) / total > 0.7:
            if 'percent' in name_lower or 'rate' in name_lower or '%' in name_lower:
                return 'NUMERIC(10,4)'
            return 'DOUBLE PRECISION'

        max_len = max(len(str(v)) for v in clean[:25])
        if max_len <= 100:
            return 'VARCHAR(255)'
        return 'TEXT'

    def _is_unit_label(self, val: str) -> bool:
        """Check if value is a unit label."""
        val_lower = val.lower().strip()
        if re.match(r'^[£$€]\d+$', val):
            return True
        unit_patterns = [
            r'^%$', r'^percent(age)?$', r'^rate$',
            r'^number$', r'^count$', r'^total$',
            r'^fte$', r'^wte$', r'^000s?$',
        ]
        for pattern in unit_patterns:
            if re.match(pattern, val_lower):
                return True
        return False

    def _is_real_numeric_data(self, val: str) -> bool:
        """Check if value is real numeric data (not a year or unit)."""
        val_clean = val.strip()
        if self.PATTERNS['calendar_year'].match(val_clean):
            return False
        if self.PATTERNS['fiscal_year'].match(val_clean.rstrip('²³¹')):
            return False
        if self._is_unit_label(val_clean):
            return False
        try:
            float(val_clean.replace(',', '').replace('£', '').replace('$', '').replace('%', ''))
            return True
        except ValueError:
            return False

    def _to_db_identifier(self, name: str) -> str:
        """Convert name to valid database identifier."""
        clean = name.lower()
        clean = re.sub(r'[£$€%]', '', clean)
        clean = re.sub(r'[^a-z0-9]+', '_', clean)
        clean = re.sub(r'_+', '_', clean).strip('_')
        if not clean:
            return 'col_unnamed'
        reserved = {'month', 'year', 'group', 'order', 'table', 'index', 'key',
                    'value', 'date', 'time', 'user', 'name', 'type', 'level'}
        if clean in reserved:
            clean = f"{clean}_val"
        if clean and clean[0].isdigit():
            clean = f"col_{clean}"
        return clean[:63]

    def _is_data_row(self, row_num: int) -> bool:
        """Check if row contains actual data."""
        numeric_count = 0
        total = 0
        for col in range(1, min(20, self.ws.max_column + 1)):
            val = self.ws.cell(row=row_num, column=col).value
            if val is not None:
                total += 1
                s = str(val).strip()
                if s.lower() not in self.SUPPRESSED_VALUES:
                    try:
                        float(s.replace(',', '').replace('£', '').replace('$', '').replace('%', ''))
                        numeric_count += 1
                    except ValueError:
                        pass
        return numeric_count >= 2 or (total >= 3 and numeric_count >= 1)

    def _find_data_end(self, data_start_row: int) -> int:
        """Find last row of data."""
        data_end = data_start_row
        empty_streak = 0
        max_row_to_scan = min(self.ws.max_row, 10000)
        max_col = min(5, self.ws.max_column)

        for row in self.ws.iter_rows(min_row=data_start_row,
                                      max_row=max_row_to_scan,
                                      min_col=1, max_col=max_col):
            has_content = False
            is_footer = False

            for cell in row:
                val = cell.value
                if val:
                    has_content = True
                    val_str = str(val).strip().lower()
                    if any(val_str.startswith(sw) for sw in self.STOP_WORDS):
                        is_footer = True
                        break

            if is_footer:
                break

            if has_content:
                data_end = row[0].row
                empty_streak = 0
            else:
                empty_streak += 1
                if empty_streak >= 5:
                    break

        return data_end

    def _count_cells(self, row_num: int) -> int:
        """Count non-empty cells in a row."""
        return sum(
            1 for col in range(1, min(50, self.ws.max_column + 1))
            if self.ws.cell(row=row_num, column=col).value is not None
        )

    def extract_data(self) -> List[Dict[str, Any]]:
        """Extract data as list of dictionaries."""
        structure = self.infer_structure()

        if not structure.is_valid:
            return []

        rows = []
        first_header_row = structure.header_rows[0] if structure.header_rows else None

        for row_num in range(structure.data_start_row, structure.data_end_row + 1):
            has_content = any(
                self.ws.cell(row=row_num, column=col).value is not None
                for col in list(structure.columns.keys())[:5]
            )

            if not has_content:
                continue

            # Check for footer
            is_footer = False
            for col_idx in list(structure.columns.keys())[:5]:
                val = self.ws.cell(row=row_num, column=col_idx).value
                if val:
                    val_str = str(val).strip().lower()
                    if any(val_str.startswith(sw) for sw in self.STOP_WORDS):
                        is_footer = True
                        break
                    if val_str.startswith('*') and len(val_str) > 50:
                        is_footer = True
                        break

            if is_footer:
                break

            # Skip duplicate header rows (section separators)
            if first_header_row:
                matches = 0
                for col_idx in list(structure.columns.keys())[:3]:
                    current_val = str(self.ws.cell(row=row_num, column=col_idx).value or '').strip().lower()
                    header_val = str(self.ws.cell(row=first_header_row, column=col_idx).value or '').strip().lower()
                    current_clean = re.sub(r'[¹²³⁴⁵⁶⁷⁸⁹⁰,]+', '', current_val)
                    header_clean = re.sub(r'[¹²³⁴⁵⁶⁷⁸⁹⁰,]+', '', header_val)
                    if current_clean and header_clean and current_clean == header_clean:
                        matches += 1
                if matches >= 2:
                    continue

            row_data = {}
            for col_idx, col_info in structure.columns.items():
                cell_val = self.ws.cell(row=row_num, column=col_idx).value
                if cell_val is not None:
                    if str(cell_val).strip().lower() in self.SUPPRESSED_VALUES:
                        cell_val = None
                row_data[col_info.pg_name] = cell_val

            rows.append(row_data)

        return rows

    def to_dataframe(self):
        """Convert extracted data to pandas DataFrame."""
        try:
            import pandas as pd
            return pd.DataFrame(self.extract_data())
        except ImportError:
            raise ImportError("pandas is required for to_dataframe()")


def get_sheet_names(filepath: str) -> List[str]:
    """Get list of sheet names from an Excel file."""
    wb = _get_cached_workbook(filepath)
    return wb.sheetnames
